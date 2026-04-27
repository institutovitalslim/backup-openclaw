#!/usr/bin/env python3
"""
Document Excel Creator
Cria planilhas .xlsx formatadas com estilos profissionais.
"""

import argparse
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_spreadsheet(output_path, sheet_name, headers, rows, currency_cols=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name or "Planilha"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Cabeçalhos
    if headers:
        headers_list = [h.strip() for h in headers.split(',')]
        for col_idx, header in enumerate(headers_list, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    
    # Dados
    if rows:
        rows_list = rows.split('\n')
        for row_idx, row_data in enumerate(rows_list, 2):
            cells = row_data.split(',')
            for col_idx, cell_value in enumerate(cells, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value.strip())
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                
                # Formatação de moeda
                if currency_cols and headers:
                    headers_list = [h.strip() for h in headers.split(',')]
                    if col_idx <= len(headers_list):
                        col_name = headers_list[col_idx - 1]
                        if col_name in [c.strip() for c in currency_cols.split(',')]:
                            try:
                                cell.number_format = 'R$ #,##0.00'
                                cell.value = float(cell_value.strip())
                            except ValueError:
                                pass
    
    # Ajustar largura das colunas
    if headers:
        headers_list = [h.strip() for h in headers.split(',')]
        for col_idx in range(1, len(headers_list) + 1):
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Congelar painel superior
    ws.freeze_panes = 'A2'
    
    # Salvar
    wb.save(output_path)
    print(f"✅ Planilha criada: {output_path}")
    return output_path

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Criar planilha Excel')
    parser.add_argument('--output', '-o', required=True, help='Caminho de saída .xlsx')
    parser.add_argument('--sheet', '-s', default='Planilha', help='Nome da aba')
    parser.add_argument('--headers', help='Cabeçalhos (separados por vírgula)')
    parser.add_argument('--rows', '-r', help='Linhas de dados (separadas por \\n, valores por vírgula)')
    parser.add_argument('--currency', '-c', help='Colunas com formato de moeda (separadas por vírgula)')
    
    args = parser.parse_args()
    
    try:
        create_spreadsheet(args.output, args.sheet, args.headers, args.rows, args.currency)
    except Exception as e:
        print(f"❌ Erro: {e}", file=sys.stderr)
        sys.exit(1)
